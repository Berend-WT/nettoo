#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genereert per race-vraagset een Excel-bestand uit de centrale vragenbank."""
from __future__ import annotations

import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent))

BANK = Path(__file__).parent / '1000+ vragen netjes gecategoriseerd.xlsx'
OUT_DIR = Path(__file__).parent / 'sets'
OUT_DIR.mkdir(exist_ok=True)

REGIONS = {
    'USA': ['Verenigde Staten', 'Amerika', 'VS ', 'VS-', 'New York', 'Washington', 'Hollywood',
            'NBA', 'NFL', 'MLB', 'Route 66', 'Silicon Valley', 'Grand Canyon', 'Florida', 'Texas',
            'Californi', 'Alaska', 'Hawaii', 'Las Vegas', 'Chicago', 'NASA', 'Mount Rushmore',
            'Yellowstone', 'Vrijheidsbeeld', 'Empire State', 'Golden Gate', 'Witte Huis',
            'Baseball', 'Basketbal', 'American football', 'McDonald', 'Coca-Cola', 'Elvis',
            'Disney', 'Super Bowl', 'touchdown', 'honkbal', 'Broadway', 'Niagarawatervallen',
            'S&P 500', 'Billboard', 'Monopoly', 'LEGO', 'KFC', 'Starbucks', 'Elvis Presley',
            'Michael Jackson', 'Taylor Swift', 'Forrest Gump', 'Tom Brady'],
    'NL': ['Nederland', 'Nederlands', 'Nederlandse', 'Schiphol', 'Amsterdam', 'Rotterdam',
           'Utrecht', 'Den Haag', 'Domtoren', 'Euromast', 'Deltawerken', 'VOC', 'Tweede Kamer',
           'Kinderdijk', 'Willem', 'Koningsdag', 'Oranje', 'Grachten', 'polder', 'Afsluitdijk',
           'Friesland', 'Flevoland', 'Rijksmuseum', 'Van Gogh', 'Rembrandt', 'Efteling',
           'stroopwafel', 'Sinterklaas', 'Elfstedentocht', 'Randstad', 'Waddeneiland',
           'Zaanse Schans', 'Cruyff', 'Cruijff', 'Van Persie', 'Ajax', 'korfbal', 'Kaasmarkt',
           'Westerscheldetunnel', 'Zeelandbrug', 'Noord-Zuidlijn', 'Máxima', 'Alkmaar', 'Gouda',
           'Edam', 'Volendam', 'Giethoorn', 'Hoge Veluwe', 'Keukenhof', 'Bollenstreek', 'Texel',
           'Maastricht', 'PSV', 'Feyenoord', 'De Kuip', 'Johan Cruijff ArenA', 'Dom van Utrecht',
           'Utrechtse'],
    'EU': ['EU-', 'Europese Unie', 'euro', 'Eurozone', 'Schengen', 'NAVO', 'Brexit', 'Londen',
           'Parijs', 'Berlijn', 'Rome', 'Romein', 'Pompeii', 'Vesuvius', 'Colosseum',
           'Eiffeltoren', 'Big Ben', 'Tower Bridge', 'Notre-Dame', 'Sagrada', 'Parthenon',
           'Acropolis', 'Matterhorn', 'Mont Blanc', 'Donau', 'Rijn', 'Seine', 'Thames', 'Alpen',
           'Pyrene', 'Spanje', 'Itali', 'Frankrijk', 'Duitsland', 'Belgi', 'Oostenrijk',
           'Zwitserland', 'Griekenland', 'Polen', 'Zweden', 'Noorwegen', 'Denemarken', 'Finland',
           'IJsland', 'Ierland', 'Portugal', 'Tsjechi', 'Hongarije', 'Roemeni', 'Kroati',
           'Middellandse', 'Noordzee', 'wereldoorlog', 'Napoleon', 'Zeus', 'Olympi', 'Athene',
           'Viking', 'Sixtijnse', 'Mona Lisa', 'Versailles', 'Stonehenge', 'Toren van Pisa',
           'Wimbledon', 'Tour de France', 'Eurovision', 'Shakespeare', 'Harry Potter',
           'Lord of the Rings', 'James Bond', 'Beethoven', 'Mozart', 'Bach', 'Picasso',
           'Monet', 'Berlijnse Muur', 'Ferrari', 'Porsche', 'BMW', 'Mercedes', 'Volkswagen',
           'Audi', 'IKEA', 'H&M', 'Zara', 'Guinness', 'Heineken', 'Carlsberg',
           'Champions League', 'Premier League', 'La Liga', 'Serie A', 'Bundesliga', 'FA Cup',
           'Nobel', 'Oscar', 'Ten Oorlog', 'kerk', 'kathedraal', 'klooster'],
    'AZIE': ['Azi', 'China', 'Chinese', 'Japan', 'Japa', 'Korea', 'India', 'Vietnam', 'Thailand',
             'Indonesi', 'Filipij', 'Maleisi', 'Singapore', 'Myanmar', 'Cambodja', 'Mongoli',
             'Kazachstan', 'Pakistan', 'Nepal', 'Sri Lanka', 'Taiwan', 'Hong Kong', 'Everest',
             'Himalaya', 'Grote Muur', 'Taj Mahal', 'Tokyo', 'Seoul', 'Beijing', 'Peking',
             'Shanghai', 'Kyoto', 'Mumbai', 'Delhi', 'Bangkok', 'Boeddha', 'Boeddhis', 'Hindoe',
             'Shinto', 'islam', 'Koran', 'Saoedi', 'Dubai', 'Isra', 'Jordani', 'Syri', 'Irak',
             'Iran', 'Turkije', 'Istanboel', 'Petra', 'Mekka', 'Ilias', 'Odyssee',
             'Alexander de Grote', 'Babylon', 'Zijderoute', 'Samurai', 'Ninja', 'Sushi',
             'Ramadan', 'Mount Fuji', 'Fuji', 'Sony', 'Samsung', 'Toyota', 'Honda', 'Shinkansen',
             'anime', 'manga', 'Nintendo', 'karate', 'kung fu', 'yoga', 'Bollywood', 'tandoori',
             'pad thai', 'nasi', 'satay', 'dim sum', 'ramen', 'sashimi', 'matcha', 'wasabi',
             'Panda', 'beverrat', 'bamboe'],
    'AFRIKA': ['Afrika', 'Egypte', 'Egyptisch', 'Marokko', 'Tunesi', 'Algerij', 'Libi', 'Soedan',
               'Ethiopi', 'Kenia', 'Tanzania', 'Oeganda', 'Rwanda', 'Nigeria', 'Ghana', 'Senegal',
               'Mali', 'Kameroen', 'Congo', 'Angola', 'Zambia', 'Zimbabwe', 'Zuid-Afrika',
               'Namibi', 'Botswana', 'Mozambique', 'Madagascar', 'Kaapstad', 'Johannesburg',
               'Kairo', 'Casablanca', 'Marrakech', 'Dakar', 'Lagos', 'Nijl', 'Sahara', 'Kalahari',
               'Serengeti', 'Sahel', 'Victoriawatervallen', 'Pyramide', 'Sfinx', 'Toetanchamon',
               'Cleopatra', 'Mandela', 'Big Five', 'Kruger', 'Zanzibar', 'Maasai', 'Farao',
               'Carthago', 'Hannibal', 'Timboektoe', 'Ngorongoro', 'Atlasgebergte', 'kora',
               'djembé'],
    'OCEANIE': ['Australi', 'Nieuw-Zeeland', 'Oceani', 'Sydney', 'Melbourne', 'Canberra', 'Perth',
                'Brisbane', 'Auckland', 'Wellington', 'Maori', 'Aboriginal', 'Uluru',
                'Ayers Rock', 'Great Barrier Reef', 'Groot Barri', 'kangoeroe', 'koala',
                'Tasmani', 'Fiji', 'Papoea', 'Nieuw-Guinea', 'Samoa', 'Tonga', 'Vanuatu',
                'Polynes', 'didgeridoo', 'Sydney Opera', 'Sydney Harbour', 'wallaby', 'emoe',
                'vogelbekdier', 'Haka', 'All Blacks', 'rugby', 'AC/DC', 'Midnight Oil', 'Mad Max',
                'Q1 Tower', 'Tour Down Under', 'Australische Open', 'Kiwis'],
    'LATIJNS_AMERIKA': ['Brazili', 'Argentin', 'Mexico', 'Chili', 'Peru', 'Colombia', 'Venezuela',
                        'Bolivia', 'Ecuador', 'Uruguay', 'Paraguay', 'Guyana', 'Suriname',
                        'Panama', 'Costa Rica', 'Guatemala', 'Cuba', 'Jamaica', 'Carib',
                        'Puerto Rico', 'Amazone', 'Amazon', 'Andes', 'Machu Picchu', 'Inca',
                        'Azteek', 'Maya', 'Chichen Itza', 'Rio de Janeiro', 'Sao Paulo',
                        'Buenos Aires', 'Santiago', 'Lima', 'Bogota', 'Caracas', 'Havana',
                        'Panamakanaal', 'tango', 'samba', 'carnaval', 'Patagoni', 'Vuurland',
                        'Galapagos', 'Atacama', 'Titicaca', 'Angel Falls', 'Salto Angel',
                        'Che Guevara', 'Yucatan', 'mariachi', 'mate', 'Copa Am', 'Mercosur',
                        'CELAC', 'Christus Redentor', 'Aconcagua', 'jaguar', 'anaconda'],
}

THEMAS = {
    'RUIMTE': ['ruimte', 'planeet', 'maan', 'zon ', 'ster ', 'melkweg', 'heelal', 'satelliet',
               'raket', 'astronaut', 'ISS', 'NASA', 'Apollo', 'Mars', 'Jupiter', 'Saturnus',
               'Venus', 'Mercurius', 'Neptunus', 'Uranus', 'Pluto', 'zwart gat', 'lichtjaar',
               'komeet', 'meteoor', 'zonnestelsel', 'ruimtestation', 'spaceshuttle', 'sterren'],
    'DIEREN': ['dier', 'hond', 'kat ', 'paard', 'koe', 'vogel', 'vis ', 'insect', 'bij ', 'mier',
               'olifant', 'leeuw', 'tijger', 'beer ', 'walvis', 'haai', 'dolfijn', 'aap',
               'schaap', 'geit', 'kip', 'eend', 'gans', 'kameleon', 'krokodil', 'slang', 'spin',
               'vlinder', 'bijenvolk', 'honing', 'poot', 'staart', 'snorharen', 'slurf', 'bulten',
               'maag', 'gewei', 'hoorn', 'schildpad', 'kameel', 'nijlpaard', 'neushoorn', 'zebra',
               'giraffe', 'luipaard', 'jachtluipaard', 'cheeta', 'hyena', 'wombat', 'platypus',
               'otter', 'bever', 'eekhoorn', 'egel', 'vleermuis', 'panda', 'gorilla', 'chimp',
               'orangutan'],
    'SPORT': ['voetbal', 'basketbal', 'honkbal', 'American football', 'tennis', 'golf', 'hockey',
              'ijshockey', 'cricket', 'rugby', 'volleybal', 'handbal', 'korfbal', 'zwemmen',
              'atletiek', 'sprint', 'marathon', 'hordeloop', 'verspringen', 'hoogspringen',
              'gewichtheffen', 'boksen', 'judo', 'karate', 'worstelen', 'schermen',
              'boogschieten', 'roeien', 'kano', 'kajak', 'zeilen', 'surfen', 'duiken', 'ski',
              'snowboard', 'schaats', 'shorttrack', 'turnen', 'gymnastiek', 'wielrennen',
              'Tour de France', 'Giro', 'Vuelta', 'Formule 1', 'F1', 'rally', 'MotoGP',
              'Olympische Spelen', 'WK', 'EK', 'Champions League', 'Wereldbeker', 'Super Bowl',
              'Wimbledon', 'US Open', 'Roland Garros', 'Australian Open'],
    'POPCULTUUR': ['film', 'serie', 'acteur', 'actrice', 'regisseur', 'Oscar', 'Golden Globe',
                   'Emmy', 'Grammy', 'Billboard', 'concert', 'tournee', 'album', 'single',
                   'nummer 1', 'hitparade', 'Netflix', 'HBO', 'Disney', 'Pixar', 'Marvel',
                   'Star Wars', 'Harry Potter', 'Lord of the Rings', 'Game of Thrones',
                   'Breaking Bad', 'Friends', 'The Office', 'Stranger Things', 'Squid Game',
                   'TikTok', 'Instagram', 'YouTube', 'Twitter', 'Facebook', 'Snapchat',
                   'WhatsApp', 'Twitch', 'streaming', 'podcast', 'videogame', 'PlayStation',
                   'Xbox', 'Nintendo', 'Minecraft', 'Fortnite', 'GTA', 'FIFA', 'Call of Duty',
                   'Super Mario', 'Zelda', 'Pok', 'Tetris', 'Pac-Man', 'Sonic', 'Roblox',
                   'Among Us', 'Candy Crush', 'Angry Birds', 'LEGO', 'Monopoly', 'Scrabble'],
}


def strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def classify(q: str) -> list[str]:
    ql = strip_accents(q).lower()
    hits = []
    for region, kws in REGIONS.items():
        for k in kws:
            if strip_accents(k).lower() in ql:
                hits.append(region)
                break
    return hits


def theme_of(q: str) -> str | None:
    ql = strip_accents(q).lower()
    best, best_n = None, 0
    for theme, kws in THEMAS.items():
        n = sum(1 for k in kws if k in ql)
        if n > best_n:
            best, best_n = theme, n
    return best


def load_rows():
    wb = load_workbook(BANK, read_only=True, data_only=True)
    ws = wb['Alle vragen']
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    i = {x: n for n, x in enumerate(h)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[i['Vraag NL']]:
            continue
        try:
            ans = int(row[i['Antwoord']])
        except (TypeError, ValueError):
            continue
        rows.append({
            'cat': str(row[i['Categorie']] or ''),
            'q': str(row[i['Vraag NL']]),
            'ans': ans,
        })
    return rows


def write_set(key: str, label: str, selected: list[dict]) -> None:
    out = Workbook()
    ws = out.active
    ws.title = 'Alle vragen'
    ws.append(['Categorie', 'Vraag NL', 'Antwoord', 'Bron EN', 'Status'])
    for r in selected:
        ws.append([r['cat'], r['q'], r['ans'], None, f'set-{key}'])
    ws.freeze_panes = 'A2'
    path = OUT_DIR / f'set_{key}.xlsx'
    out.save(path)
    print(f'{key:22s} {len(selected):4d} vragen -> {path.name} ({label})')


def main():
    rows = load_rows()
    print('Vragenbank:', len(rows))

    # Volledige regio-hits per vraag (kan er meerdere zijn)
    hits_map = {r['q']: classify(r['q']) for r in rows}

    def region_in(region):
        return lambda q, c: region in hits_map[q]

    defs = [
        ('standaard', 'Wereldset (exclusief echt NL-vragen)',
         lambda q, c: 'NL' not in hits_map[q]),
        ('nederland', 'Nederland',
         lambda q, c: 'NL' in hits_map[q]),
        ('usa', 'Verenigde Staten', region_in('USA')),
        ('europa', 'Europa', region_in('EU')),
        ('azie', 'Azië', region_in('AZIE')),
        ('afrika', 'Afrika', region_in('AFRIKA')),
        ('oceanie', 'Oceanië', region_in('OCEANIE')),
        ('latijns_amerika', 'Latijns-Amerika', region_in('LATIJNS_AMERIKA')),
        ('ruimte_wetenschap', 'Ruimte & Wetenschap',
         lambda q, c: theme_of(q) == 'RUIMTE' or c in ('Sterrenkunde & ruimte', 'Natuurkunde',
                                                       'Scheikunde', 'Wetenschap', 'Technologie')),
        ('dierenrijk', 'Dierenrijk',
         lambda q, c: theme_of(q) == 'DIEREN' or c in ('Dieren', 'Biologie & gezondheid')),
        ('sport', 'Sport', lambda q, c: theme_of(q) == 'SPORT' or c == 'Sport'),
        ('popcultuur', 'Popcultuur',
         lambda q, c: theme_of(q) == 'POPCULTUUR' or c in ('Films en series', 'Muziek',
                                                           'Spellen en speelgoed',
                                                           'Merken en producten',
                                                           'Mode en lifestyle')),
    ]
    for key, label, pred in defs:
        selected = [r for r in rows if pred(r['q'], r['cat'])]
        write_set(key, label, selected)


if __name__ == '__main__':
    main()
