#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Voegt extra setvragen toe aan de hoofdbank én de set-bestanden (met dedup)."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from extra_set_vragen import EXTRA

BANK = Path(__file__).parent / '1000+ vragen netjes gecategoriseerd.xlsx'
SETS_DIR = Path(__file__).parent / 'sets'


def main():
    wb = load_workbook(BANK)
    ws = wb['Alle vragen']
    headers = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}

    existing = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col['Vraag NL']).value
        if v:
            existing.add(str(v).strip().lower())

    fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    added_total = 0
    for set_key, questions in EXTRA.items():
        set_path = SETS_DIR / f'set_{set_key}.xlsx'
        swb = load_workbook(set_path)
        sws = swb['Alle vragen']
        added = 0
        for cat, question, answer in questions:
            key = question.strip().lower()
            if key in existing:
                continue
            r = ws.max_row + 1
            ws.cell(r, col['Categorie']).value = cat
            ws.cell(r, col['Vraag NL']).value = question
            ws.cell(r, col['Antwoord']).value = answer
            ws.cell(r, col['Status']).value = f'nieuw-2026-09-{set_key}'
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = fill

            sr = sws.max_row + 1
            sws.cell(sr, 1).value = cat
            sws.cell(sr, 2).value = question
            sws.cell(sr, 3).value = answer
            sws.cell(sr, 5).value = f'set-{set_key}'
            existing.add(key)
            added += 1
        swb.save(set_path)
        added_total += added
        print(f'{set_key}: +{added} vragen')

    wb.save(BANK)
    print('Totaal toegevoegd:', added_total)
    print('Bank nu:', ws.max_row - 1, 'vragen')


if __name__ == '__main__':
    main()
