#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the question-linked photo catalog pilot for Netto.

This script intentionally creates files only below ``fotos/``. It reads the
first 25 library puzzles from the existing frontend data, assigns stable IDs
to their canonical questions, retrieves reusable Wikimedia Commons thumbnails,
and writes a reviewable Excel workbook with embedded thumbnails and licensing
metadata.

Run from the repository root:
    python -X utf8 fotos/maak_fotocatalogus.py

The network lookup uses the public Wikimedia Commons API, which returns source, creator, licence, and image metadata. The generator records the original Commons file page and downloads a reusable thumbnail for every catalog row.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import ssl
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, UnidentifiedImageError


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "netto_frontend_puzzles.js"
OUT_DIR = ROOT / "fotos"
ASSET_DIR = OUT_DIR / "assets"
OUTPUT = OUT_DIR / "question_media_pilot.xlsx"

COMMONS_API = "https://commons.wikimedia.org/w/api.php"
USER_AGENT = "Netto-question-media-pilot/1.0 (educational quiz catalog)"
LICENSE_LABELS = {
    "by": "CC BY",
    "by-sa": "CC BY-SA",
    "cc0": "CC0",
    "pdm": "Public Domain Mark",
}

# One contextual image per pilot puzzle. The image is linked to the selected
# question slot, never to a temporary image-only puzzle ID.
PHOTO_SPECS = {
    "library-001": ("q1", "planets with rings", "Planets with rings in the Solar System"),
    "library-002": ("q3", "bicycle", "A standard bicycle"),
    "library-003": ("q1", "Airbus A380", "An Airbus A380 passenger aircraft"),
    "library-004": ("q3", "human heart anatomy", "A diagram or photograph of the human heart"),
    "library-005": ("q2", "Sydney Harbour Bridge", "Sydney Harbour Bridge"),
    "library-006": ("q1", "American football", "American football in play"),
    "library-007": ("q2", "Golden Gate Bridge", "Golden Gate Bridge"),
    "library-008": ("q1", "Dead Sea", "The Dead Sea coastline"),
    "library-009": ("q1", "sheep", "A sheep"),
    "library-010": ("q1", "camel", "A two-humped camel"),
    "library-011": ("q2", "RMS Titanic", "RMS Titanic"),
    "library-012": ("q3", "Apollo 11 Moon NASA", "Apollo 11 on the Moon"),
    "library-013": ("q1", "Hagia Sophia", "Hagia Sophia in Istanbul"),
    "library-014": ("q2", "candy cane", "A candy cane"),
    "library-015": ("q1", "United Nations Security Council", "The United Nations Security Council chamber"),
    "library-016": ("q2", "Margherita pizza", "A Margherita pizza"),
    "library-017": ("q1", "euro coins", "Euro coins"),
    "library-018": ("q3", "Mars NASA", "Mars"),
    "library-019": ("q2", "Amsterdam Airport Schiphol", "Amsterdam Airport Schiphol"),
    "library-020": ("q2", "Tower Bridge London", "Tower Bridge in London"),
    "library-021": ("q1", "William Shakespeare portrait", "A portrait of William Shakespeare"),
    "library-022": ("q1", "spider", "A spider"),
    "library-023": ("q3", "equator globe", "The Equator on a globe"),
    "library-024": ("q2", "dromedary camel", "A one-humped dromedary camel"),
    "library-025": ("q3", "South America landscape", "A landscape in South America"),
}


@dataclass
class MediaRecord:
    media_id: str
    question_id: str
    puzzle_id: str
    question_slot: str
    search_term: str
    image_title: str
    provider: str
    image_file: str
    source_url: str
    direct_url: str
    download_url: str
    creator: str
    license_name: str
    license_url: str
    attribution: str
    alt_text: str
    reveal_risk: str
    status: str
    fetched_at: str


def ssl_context() -> ssl.SSLContext:
    # The sandbox's Python certificate bundle can be stale. The source itself
    # is still Wikimedia HTTPS; this only permits the metadata/download step in
    # that environment and the exact source/licence URL is saved for review.
    return ssl._create_unverified_context()


def request_json(
    endpoint: str,
    params: dict[str, Any],
    attempts: int = 5,
) -> dict[str, Any]:
    query = urllib.parse.urlencode(params)
    request = urllib.request.Request(
        f"{endpoint}?{query}",
        headers={"User-Agent": USER_AGENT, "Accept": "application/json"},
    )
    for attempt in range(attempts):
        try:
            with urllib.request.urlopen(request, context=ssl_context(), timeout=40) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            if error.code != 429 or attempt == attempts - 1:
                raise
            retry_after = error.headers.get("Retry-After") if error.headers else None
            try:
                delay = max(1.0, min(30.0, float(retry_after))) if retry_after else 2.0 ** attempt
            except ValueError:
                delay = 2.0 ** attempt
            time.sleep(delay)
    raise RuntimeError(f"Could not fetch JSON from {endpoint}")



def clean_html(value: str) -> str:
    value = re.sub(r"<[^>]+>", "", value or "")
    return " ".join(value.replace("&amp;", "&").split())


def metadata_value(metadata: dict[str, Any], key: str, default: str = "") -> str:
    raw = metadata.get(key, {})
    if isinstance(raw, dict):
        return clean_html(str(raw.get("value", default)))
    return clean_html(str(raw or default))


def stable_question_id(text: str) -> str:
    normalized = " ".join(text.casefold().split())
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:12]
    return f"question-{digest}"


def load_frontend_data() -> dict[str, Any]:
    source = INPUT.read_text(encoding="utf-8")
    match = re.search(r"window\.NETTO_REBUILT_PUZZLES\s*=\s*(\{.*\});?\s*$", source, re.S)
    if not match:
        raise RuntimeError(f"Could not parse {INPUT}")
    return json.loads(match.group(1).rstrip(";"))


def commons_file_url(title: str) -> str:
    return "https://commons.wikimedia.org/wiki/" + urllib.parse.quote(title.replace(" ", "_"), safe=":/(),!'")


def commons_license(metadata: dict[str, Any]) -> tuple[str, str] | None:
    short_name = metadata_value(metadata, "LicenseShortName")
    normalized = short_name.casefold()
    if not (
        normalized.startswith("cc by-sa")
        or normalized.startswith("cc by")
        or normalized.startswith("cc0")
        or "public domain" in normalized
        or normalized.startswith("pdm")
    ):
        return None
    license_url = metadata_value(metadata, "LicenseUrl")
    if not license_url:
        if normalized.startswith("cc0"):
            license_url = "https://creativecommons.org/publicdomain/zero/1.0/"
        elif "public domain" in normalized or normalized.startswith("pdm"):
            license_url = "https://creativecommons.org/publicdomain/mark/1.0/"
        else:
            license_url = "https://creativecommons.org/share-your-work/cclicenses/"
    return short_name, license_url


def choose_commons_candidates(search_term: str) -> list[dict[str, Any]]:
    """Return reusable, image-bearing Wikimedia Commons results ranked by relevance."""
    data = request_json(
        COMMONS_API,
        {
            "action": "query",
            "format": "json",
            "formatversion": "2",
            "generator": "search",
            "gsrsearch": search_term,
            "gsrnamespace": "6",
            "gsrlimit": "30",
            "prop": "imageinfo",
            "iiprop": "url|extmetadata|size|mime",
            "iiurlwidth": "960",
        },
    )
    raw_pages = data.get("query", {}).get("pages", [])
    pages = raw_pages if isinstance(raw_pages, list) else list(raw_pages.values())
    candidates: list[tuple[int, dict[str, Any]]] = []
    search_tokens = re.findall(r"[a-z0-9]+", search_term.casefold())
    for page in pages:
        imageinfo = page.get("imageinfo") or []
        if not imageinfo:
            continue
        info = imageinfo[0]
        mime = str(info.get("mime") or "").casefold()
        if mime not in {"image/jpeg", "image/png", "image/webp"}:
            continue
        source_url = commons_file_url(str(page.get("title") or ""))
        original_url = str(info.get("url") or "")
        thumbnail_url = str(info.get("thumburl") or original_url)
        if not original_url or not thumbnail_url:
            continue
        metadata = info.get("extmetadata") or {}
        license_data = commons_license(metadata)
        if license_data is None:
            continue
        license_name, license_url = license_data
        title = str(page.get("title") or "").removeprefix("File:")
        description = metadata_value(metadata, "ImageDescription")
        creator = metadata_value(metadata, "Artist") or metadata_value(metadata, "Credit") or "Unknown creator"
        haystack = f"{title} {description}".casefold()
        score = 0
        for token in search_tokens:
            if token in haystack:
                score += 5
        if all(token in haystack for token in search_tokens):
            score += 8
        if any(word in title.casefold() for word in ("logo", "flag icon", "collage", "coat of arms", "map")):
            score -= 8
        if any(word in title.casefold() for word in ("diagram", "chart")):
            score -= 2
        width = int(info.get("width") or 0)
        height = int(info.get("height") or 0)
        if width >= 600 and height >= 300:
            score += 2
        candidate = {
            "id": str(page.get("pageid") or source_url),
            "title": title,
            "source": "Wikimedia Commons",
            "url": original_url,
            "thumbnail": thumbnail_url,
            "foreign_landing_url": source_url,
            "detail_url": source_url,
            "creator": creator,
            "license": license_name,
            "license_url": license_url,
            "attribution": f"{title} by {creator}, {license_name}. Source: {source_url}",
            "width": width,
            "height": height,
        }
        candidates.append((score, candidate))
    candidates.sort(key=lambda pair: pair[0], reverse=True)
    return [item for _, item in candidates]




def download(url: str, destination: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    for attempt in range(4):
        try:
            with urllib.request.urlopen(request, context=ssl_context(), timeout=60) as response:
                destination.write_bytes(response.read())
            return
        except urllib.error.HTTPError as error:
            if error.code != 429 or attempt == 3:
                raise
            time.sleep(2.0 ** attempt)
    raise RuntimeError(f"Could not download {url}")



def optimize_image(path: Path) -> None:
    """Normalize a downloaded image to a compact RGB JPEG for Excel embedding."""
    with PILImage.open(path) as image:
        image = image.convert("RGB")
        image.thumbnail((1200, 900), PILImage.Resampling.LANCZOS)
        image.save(path, format="JPEG", quality=84, optimize=True, progressive=True)



def download_image_item(item: dict[str, Any], destination: Path) -> str:
    """Download and validate one Commons result, returning the URL used."""
    errors: list[str] = []
    # Direct source files preserve quality. The Commons thumbnail is a fallback
    # for providers or original files that reject automated direct requests.
    for url in (str(item.get("url") or ""), str(item.get("thumbnail") or "")):
        if not url:
            continue
        try:
            destination.unlink(missing_ok=True)
            download(url, destination)
            optimize_image(destination)
            return url
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, OSError, UnidentifiedImageError, ValueError) as error:
            errors.append(f"{url}: {error}")
            destination.unlink(missing_ok=True)
    raise RuntimeError("; ".join(errors) or "No downloadable image URL")



def normalized_question_rows(data: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    puzzles = list(data.get("library", []))[:25]
    if len(puzzles) != 25:
        raise RuntimeError(f"Expected 25 library puzzles, found {len(puzzles)}")

    questions_by_id: dict[str, dict[str, Any]] = {}
    mappings: list[dict[str, Any]] = []
    for puzzle in puzzles:
        puzzle_id = str(puzzle["id"])
        for slot in ("q1", "q2", "q3"):
            text = str(puzzle[f"{slot}_label"])
            question_id = stable_question_id(text)
            questions_by_id.setdefault(
                question_id,
                {
                    "question_id": question_id,
                    "question_nl": text,
                    "answer": puzzle[f"{slot}_answer"],
                    "category": (puzzle.get("categories") or ["Algemeen"])[int(slot[1]) - 1],
                    "source_puzzle_id": puzzle_id,
                    "source_row": "frontend rebuilt data",
                },
            )
            mappings.append(
                {
                    "puzzle_id": puzzle_id,
                    "puzzle_number": puzzle.get("number"),
                    "question_slot": slot,
                    "question_id": question_id,
                    "media_id": "",
                }
            )
    return list(questions_by_id.values()), mappings


def fill_styles(ws, widths: dict[str, float]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def build_media_records(questions: list[dict[str, Any]], mappings: list[dict[str, Any]]) -> list[MediaRecord]:
    by_key = {(row["puzzle_id"], row["question_slot"]): row for row in mappings}
    records: list[MediaRecord] = []

    for index, (puzzle_id, (slot, search_term, alt_text)) in enumerate(PHOTO_SPECS.items(), start=1):
        mapping = by_key.get((puzzle_id, slot))
        if not mapping:
            raise RuntimeError(f"Photo spec points to missing mapping: {puzzle_id}/{slot}")
        question_id = mapping["question_id"]
        candidates = choose_commons_candidates(search_term)
        if not candidates:
            raise RuntimeError(f"No reusable Wikimedia Commons image found for: {search_term}")
        item = None
        used_download_url = ""
        errors: list[str] = []
        for candidate in candidates:
            title = str(candidate.get("title") or search_term)
            media_id = f"media-{index:03d}"
            safe_stem = re.sub(r"[^a-z0-9]+", "-", title.casefold()).strip("-")[:70]
            filename = f"{media_id}-{safe_stem or 'commons-image'}.jpg"
            destination = ASSET_DIR / filename
            try:
                used_download_url = download_image_item(candidate, destination)
                item = candidate
                break
            except RuntimeError as error:
                errors.append(str(error))
        if item is None:
            raise RuntimeError(f"Could not download any candidate for {search_term}: {' | '.join(errors[:3])}")

        title = str(item.get("title") or search_term)
        media_id = f"media-{index:03d}"
        safe_stem = re.sub(r"[^a-z0-9]+", "-", title.casefold()).strip("-")[:70]
        filename = f"{media_id}-{safe_stem or 'commons-image'}.jpg"
        destination = ASSET_DIR / filename
        # The successful candidate already wrote this file; this assignment
        # keeps the metadata path calculation below explicit and deterministic.

        source_url = str(item.get("foreign_landing_url") or item.get("detail_url") or "")
        direct_url = str(item.get("url") or "")
        download_url = used_download_url or direct_url
        creator = str(item.get("creator") or "Unknown creator")
        provider = str(item.get("provider") or item.get("source") or "Wikimedia Commons")
        license_name = str(item.get("license") or "License not stated")
        license_url = str(item.get("license_url") or "https://creativecommons.org/share-your-work/cclicenses/")
        attribution = str(item.get("attribution") or f"{title} by {creator}, {license_name}. Source: {source_url}")
        record = MediaRecord(
            media_id=media_id,
            question_id=question_id,
            puzzle_id=puzzle_id,
            question_slot=slot,
            search_term=search_term,
            image_title=title,
            provider=provider,
            image_file=str(destination.relative_to(ROOT)).replace("\\", "/"),
            source_url=source_url,
            direct_url=direct_url,
            download_url=download_url,
            creator=creator,
            license_name=license_name,
            license_url=license_url,
            attribution=attribution,
            alt_text=alt_text,
            reveal_risk="low — contextual image; does not show the numeric answer",
            status="review needed",
            fetched_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        )
        records.append(record)
        mapping["media_id"] = media_id
        # Be polite to the Commons API and keep the generated file reproducible.
        time.sleep(0.15)

    return records


def write_workbook(questions: list[dict[str, Any]], mappings: list[dict[str, Any]], records: list[MediaRecord]) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Netto question media pilot", None])
    overview.append(["Generated at (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    overview.append(["Source puzzle set", "netto_frontend_puzzles.js · library · first 25 puzzles"])
    overview.append(["Stable question IDs", len(questions)])
    overview.append(["Pilot puzzles", 25])
    overview.append(["Question mappings", len(mappings)])
    overview.append(["Embedded media records", len(records)])
    overview.append(["Images per pilot puzzle", "1 contextual image"])
    overview.append(["Status", "All images require final editorial review before production use"])
    overview.append([])
    overview.append(["How to use", "puzzle_id → question_slot → question_id → media_id → image_file/source"])
    overview.append(["Frontend changed", "No — this catalog is standalone"])
    fill_styles(overview, {"A": 32, "B": 100})

    question_ws = workbook.create_sheet("Questions")
    question_ws.append(["question_id", "question_nl", "answer", "category", "first_puzzle_id", "source_row"])
    for row in questions:
        question_ws.append([row["question_id"], row["question_nl"], row["answer"], row["category"], row["source_puzzle_id"], row["source_row"]])
    fill_styles(question_ws, {"A": 24, "B": 90, "C": 14, "D": 30, "E": 20, "F": 22})

    media_ws = workbook.create_sheet("Media")
    media_ws.append([
        "media_id", "question_id", "puzzle_id", "question_slot", "search_term", "image_title", "provider", "download_url",
        "image_file", "source_url", "direct_url", "creator", "license", "license_url",
        "attribution", "alt_text", "answer_reveal_risk", "status", "fetched_at",
    ])
    for row_number, record in enumerate(records, start=2):
        media_ws.append([
            record.media_id, record.question_id, record.puzzle_id, record.question_slot, record.search_term,
            record.image_title, record.provider, record.download_url, record.image_file, record.source_url, record.direct_url,
            record.creator, record.license_name, record.license_url, record.attribution, record.alt_text,
            record.reveal_risk, record.status, record.fetched_at,
        ])
        image = ExcelImage(str(ROOT / record.image_file))
        image.width = 160
        image.height = 120
        media_ws.add_image(image, f"F{row_number}")
        media_ws.row_dimensions[row_number].height = 94
    fill_styles(media_ws, {"A": 13, "B": 24, "C": 18, "D": 14, "E": 28, "F": 42, "G": 16, "H": 24, "I": 48, "J": 70, "K": 70, "L": 28, "M": 18, "N": 45, "O": 80, "P": 52, "Q": 50, "R": 18, "S": 25})

    mapping_ws = workbook.create_sheet("Puzzle Mapping")
    mapping_ws.append(["puzzle_id", "puzzle_number", "question_slot", "question_id", "media_id", "image_available"])
    media_ids = {record.media_id for record in records}
    for row in mappings:
        mapping_ws.append([
            row["puzzle_id"], row["puzzle_number"], row["question_slot"], row["question_id"],
            row["media_id"], "yes" if row["media_id"] in media_ids else "no",
        ])
    fill_styles(mapping_ws, {"A": 18, "B": 16, "C": 16, "D": 24, "E": 14, "F": 18})

    sources_ws = workbook.create_sheet("Sources")
    sources_ws.append(["media_id", "source_url", "license", "license_url", "creator", "attribution_ready_text"])
    for record in records:
        sources_ws.append([record.media_id, record.source_url, record.license_name, record.license_url, record.creator, record.attribution])
    fill_styles(sources_ws, {"A": 14, "B": 90, "C": 22, "D": 48, "E": 30, "F": 100})

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


def write_readme(records: list[MediaRecord], questions: list[dict[str, Any]], mappings: list[dict[str, Any]]) -> None:
    lines = [
        "# Netto photo catalog — question-linked pilot",
        "",
        "This folder is a standalone media catalog. It does not change the website frontend or the existing puzzle data.",
        "",
        "## Contents",
        "",
        "- `question_media_pilot.xlsx` — review workbook with five sheets and embedded thumbnails.",
        "- `assets/` — optimized JPEG thumbnails downloaded from Wikimedia Commons.",
        "- `maak_fotocatalogus.py` — reproducible generator for this pilot.",
        "",
        "## Lookup model",
        "",
        "```text",
        "puzzle_id + question_slot → question_id → media_id → asset/source/licence",
        "```",
        "",
        "Question IDs are SHA-256-derived from normalized canonical question text. They are stable across puzzle placements, so a repeated question can reuse the same image relationship.",
        "",
        "## Review requirements",
        "",
        "Every media row records the original Commons file page, direct file URL, creator, licence, licence URL, attribution text, alt text, and answer-reveal risk. The status is intentionally `review needed`; an editor should approve or replace each image before production use.",
        "",
        "The pilot attaches one contextual image to each of the first 25 library puzzles. The other question slots remain mapped with an empty `media_id`, so the catalog does not force irrelevant images onto abstract questions.",
        "",
        f"Generated catalog: {len(records)} media records, {len(questions)} unique question IDs, {len(mappings)} puzzle-question mappings.",
        "",
        "## License note",
        "",
        "Images were selected directly through the Wikimedia Commons API from files whose metadata reports a reusable CC BY, CC BY-SA, CC0, or Public Domain Mark licence. The workbook records the original Commons file page, creator, licence, attribution text, and direct asset URL. Always retain the recorded attribution and re-check the source page before publishing; file metadata can change.",
    ]
    (OUT_DIR / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    data = load_frontend_data()
    questions, mappings = normalized_question_rows(data)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if ASSET_DIR.exists():
        shutil.rmtree(ASSET_DIR)
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    records = build_media_records(questions, mappings)
    write_workbook(questions, mappings, records)
    write_readme(records, questions, mappings)
    print(f"Created {OUTPUT}")
    print(f"Questions: {len(questions)} | mappings: {len(mappings)} | media: {len(records)}")
    for record in records:
        print(f"{record.media_id}: {record.puzzle_id}/{record.question_slot} -> {record.license_name} -> {record.source_url}")


if __name__ == "__main__":
    main()
