#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genereer de Netto-hoofdset en de puzzels voor Puzzel Race.

De hoofdset bevat exact 200 puzzels:
- 50 Easy, 50 Intermediate, 50 Hard en 50 Extremely Hard;
- 50 vermenigvuldig-, 50 deel-, 50 plus- en 50 minsommen;
- iedere vraag wordt in de hoofdset maximaal één keer gebruikt.

Na de hoofdset worden met de overgebleven vragen zoveel mogelijk extra geldige
puzzels gemaakt voor Puzzel Race. Daily Archive wordt later uit de Hard-hoofdset
opgebouwd door .freebuff/build_rebuilt_frontend.py.
"""

from __future__ import annotations

import argparse
import math
import random
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from fractions import Fraction
from pathlib import Path
from typing import DefaultDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = "1000+ vragen netjes gecategoriseerd.xlsx"
DEFAULT_OUTPUT = "netto_puzzels_race.xlsx"

LEVELS = ("Easy", "Intermediate", "Hard", "Extremely Hard")
OPERATORS = ("×", "÷", "+", "−")
OPERATOR_WEIGHTS = {"+": 0.0, "−": 8.0, "×": 14.0, "÷": 18.0}

# Elke operator krijgt 50 puzzels. De schaarse Easy-relaties krijgen een
# kleinere, maar voldoende quota; de andere levels vullen de 50 per level aan.
MAIN_QUOTAS = {
    "×": {"Easy": 5, "Intermediate": 15, "Hard": 15, "Extremely Hard": 15},
    "÷": {"Easy": 5, "Intermediate": 15, "Hard": 15, "Extremely Hard": 15},
    "+": {"Easy": 20, "Intermediate": 10, "Hard": 10, "Extremely Hard": 10},
    "−": {"Easy": 20, "Intermediate": 10, "Hard": 10, "Extremely Hard": 10},
}


@dataclass(frozen=True)
class Question:
    source_row: int
    category: str
    text: str
    answer: Fraction


@dataclass(frozen=True)
class Relation:
    operator: str
    values: tuple[Fraction, Fraction, Fraction]
    level: str
    score: float


@dataclass(frozen=True)
class Puzzle:
    operator: str
    ids: tuple[int, int, int]
    level: str
    score: float


def parse_integer(value: object) -> int | None:
    """Lees alleen hele numerieke antwoorden in; decimalen worden overgeslagen."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            return None
        return int(value)

    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        return None
    if re.fullmatch(r"[-+]?\d{1,3}(?:\.\d{3})+", text):
        text = text.replace(".", "")
    elif "," in text or "." in text:
        return None
    if not re.fullmatch(r"[-+]?\d+", text):
        return None
    return int(text)


def normalize_text(text: str) -> str:
    return " ".join(text.casefold().split())


def resolve_path(argument: str | None, default_name: str) -> Path:
    return Path(argument).expanduser() if argument else BASE_DIR / default_name


def difficulty_for(operator: str, values: tuple[Fraction, Fraction, Fraction]) -> tuple[str, float]:
    score = OPERATOR_WEIGHTS[operator] + sum(
        math.log10(abs(float(value)) + 1.0) * 10.0 for value in values
    )
    score = round(score, 2)
    if score < 35:
        return "Easy", score
    if score < 60:
        return "Intermediate", score
    if score < 90:
        return "Hard", score
    return "Extremely Hard", score


def load_questions(path: Path) -> tuple[list[Question], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Alle vragen" not in workbook.sheetnames:
        raise ValueError("Het invoerbestand bevat geen tabblad 'Alle vragen'.")

    questions: list[Question] = []
    skipped: list[str] = []
    seen: set[str] = set()
    worksheet = workbook["Alle vragen"]
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 3 or row[1] in (None, ""):
            continue
        text = str(row[1]).strip()
        answer = parse_integer(row[2])
        if answer is None:
            skipped.append(f"Rij {row_number}: geen geheel antwoord ({row[2]!r})")
            continue
        key = normalize_text(text)
        if key in seen:
            skipped.append(f"Rij {row_number}: dubbele vraagtekst")
            continue
        seen.add(key)
        questions.append(
            Question(
                source_row=row_number,
                category=str(row[0] or "Algemeen"),
                text=text,
                answer=Fraction(answer),
            )
        )
    workbook.close()
    return questions, skipped


def relation_values(operator: str, value_set: set[Fraction]) -> list[tuple[Fraction, Fraction, Fraction]]:
    values = sorted(value_set)
    relations: list[tuple[Fraction, Fraction, Fraction]] = []
    if operator == "×":
        usable = [value for value in values if value > 0 and value != 1]
        for first_position, first in enumerate(usable):
            for second in usable[first_position:]:
                result = first * second
                if result in value_set:
                    relations.append((first, second, result))
    elif operator == "÷":
        usable = [value for value in values if value > 0 and value != 1]
        for divisor in usable:
            for quotient in usable:
                dividend = divisor * quotient
                if dividend in value_set:
                    relations.append((dividend, divisor, quotient))
    elif operator == "+":
        usable = [value for value in values if value > 0]
        for first_position, first in enumerate(usable):
            for second in usable[first_position:]:
                result = first + second
                if result in value_set:
                    relations.append((first, second, result))
    elif operator == "−":
        positive = [value for value in values if value > 0]
        for first in positive:
            for second in positive:
                if second >= first:
                    continue
                result = first - second
                if result > 0 and result in value_set:
                    relations.append((first, second, result))
    return relations


def relation_counts(values: tuple[Fraction, Fraction, Fraction]) -> Counter[Fraction]:
    return Counter(values)


def fits(values: tuple[Fraction, Fraction, Fraction], capacities: dict[Fraction, int]) -> bool:
    return all(count <= capacities.get(value, 0) for value, count in relation_counts(values).items())


def build_relations(questions: list[Question]) -> dict[tuple[str, str], list[Relation]]:
    value_set = {question.answer for question in questions}
    grouped: dict[tuple[str, str], list[Relation]] = defaultdict(list)
    for operator in OPERATORS:
        for values in relation_values(operator, value_set):
            level, score = difficulty_for(operator, values)
            grouped[(operator, level)].append(Relation(operator, values, level, score))
    return grouped


def allocate_ids(
    values: tuple[Fraction, Fraction, Fraction],
    available: DefaultDict[Fraction, list[int]],
) -> tuple[int, int, int]:
    chosen: list[int] = []
    local: Counter[Fraction] = Counter()
    for value in values:
        local[value] += 1
        ids = available[value]
        if len(ids) < local[value]:
            raise RuntimeError("Interne fout: relation past niet bij de beschikbare vragen.")
        chosen.append(ids.pop())
    return chosen[0], chosen[1], chosen[2]


def group_value_degrees(relations: list[Relation]) -> Counter[Fraction]:
    degrees: Counter[Fraction] = Counter()
    for relation in relations:
        degrees.update(set(relation.values))
    return degrees


def choose_main_set(
    questions: list[Question],
    seed: int,
    max_attempts: int = 80,
) -> tuple[list[Puzzle], set[int]]:
    """Zoek met randomized scarcity-first selectie een exacte quota-oplossing."""
    all_relations = build_relations(questions)
    groups = [(operator, level) for operator in OPERATORS for level in LEVELS]
    target_total = sum(MAIN_QUOTAS[operator][level] for operator, level in groups)
    rng = random.Random(seed)

    for attempt in range(max_attempts):
        available: DefaultDict[Fraction, list[int]] = defaultdict(list)
        for index, question in enumerate(questions):
            available[question.answer].append(index)
        for ids in available.values():
            rng.shuffle(ids)

        remaining = {
            group: MAIN_QUOTAS[group[0]][group[1]] for group in groups
        }
        selected: list[Puzzle] = []
        used: set[int] = set()
        failed = False

        for _step in range(target_total):
            capacities = {value: len(ids) for value, ids in available.items()}
            candidates_by_group: dict[tuple[str, str], list[Relation]] = {}
            active_groups: list[tuple[float, int, tuple[str, str]]] = []

            for group in groups:
                needed = remaining[group]
                if needed <= 0:
                    continue
                feasible = [
                    relation
                    for relation in all_relations[group]
                    if fits(relation.values, capacities)
                ]
                if not feasible:
                    failed = True
                    break
                candidates_by_group[group] = feasible
                # Kies het meest kritieke quota eerst. Dat voorkomt dat een
                # schaarse difficulty pas aan het einde geen ruimte meer heeft.
                active_groups.append((len(feasible) / needed, len(feasible), group))
            if failed:
                break

            active_groups.sort(key=lambda item: (item[0], item[1]))
            critical_ratio = active_groups[0][0]
            tied = [item for item in active_groups if item[0] <= critical_ratio * 1.12]
            _, _, group = rng.choice(tied[: min(4, len(tied))])
            feasible = candidates_by_group[group]

            degrees = group_value_degrees(all_relations[group])
            scored: list[tuple[float, Relation]] = []
            for relation in feasible:
                counts = relation_counts(relation.values)
                scarcity = sum(count / max(1, capacities[value]) for value, count in counts.items())
                rarity = sum(count / max(1, degrees[value]) for value, count in counts.items())
                # Een kleine random component maakt nieuwe seeds werkelijk
                # verschillend zonder de schaarsteheuristiek te verliezen.
                score = scarcity * 4.0 + rarity * 2.0 + rng.random() * 0.05
                scored.append((score, relation))
            scored.sort(key=lambda item: item[0], reverse=True)
            top = scored[: min(32, len(scored))]
            relation = rng.choice(top)[1]

            ids = allocate_ids(relation.values, available)
            selected.append(Puzzle(relation.operator, ids, relation.level, relation.score))
            used.update(ids)
            remaining[group] -= 1

        if not failed and len(selected) == target_total and not any(remaining.values()):
            return selected, used

    raise RuntimeError(
        f"Geen exacte hoofdset gevonden na {max_attempts} pogingen. "
        "Probeer een andere seed of controleer de vragenbank."
    )


def choose_race_set(
    questions: list[Question],
    used: set[int],
    seed: int,
) -> tuple[list[Puzzle], set[int]]:
    """Maak extra puzzels met uitsluitend nog ongebruikte vragen."""
    rng = random.Random(seed)
    available: DefaultDict[Fraction, list[int]] = defaultdict(list)
    for index, question in enumerate(questions):
        if index not in used:
            available[question.answer].append(index)
    for ids in available.values():
        rng.shuffle(ids)

    race: list[Puzzle] = []
    race_used: set[int] = set()
    while True:
        capacities = {value: len(ids) for value, ids in available.items()}
        value_set = set(available)
        feasible: list[Relation] = []
        for operator in OPERATORS:
            for values in relation_values(operator, value_set):
                if not fits(values, capacities):
                    continue
                level, score = difficulty_for(operator, values)
                feasible.append(Relation(operator, values, level, score))
        if not feasible:
            break

        # Gebruik relaties met drie verschillende waarden bij voorkeur; dat
        # laat vaak meer vervolgcombinaties over voor de racepool.
        diverse = [relation for relation in feasible if len(set(relation.values)) == 3]
        pool = diverse or feasible
        pool.sort(key=lambda relation: (len(set(relation.values)), -relation.score), reverse=True)
        top = pool[: min(48, len(pool))]
        relation = rng.choice(top)
        ids = allocate_ids(relation.values, available)
        race.append(Puzzle(relation.operator, ids, relation.level, relation.score))
        race_used.update(ids)

    return race, race_used


def number_text(value: Fraction) -> int:
    if value.denominator != 1:
        raise ValueError("Alle antwoorden moeten gehele getallen zijn.")
    return value.numerator


def formula(operator: str, values: tuple[Fraction, Fraction, Fraction]) -> str:
    return f"{number_text(values[0])} {operator} {number_text(values[1])} = {number_text(values[2])}"


def write_puzzle_sheet(
    worksheet,
    title_column: str,
    puzzles: list[Puzzle],
    questions: list[Question],
) -> None:
    headers = [
        title_column,
        "Bewerking",
        "Antwoord 1",
        "Categorie 1",
        "Vraag 1",
        "Antwoord 2",
        "Categorie 2",
        "Vraag 2",
        "Antwoord 3",
        "Categorie 3",
        "Vraag 3",
        "Exacte formule",
        "Difficulty",
        "Difficulty score",
    ]
    worksheet.append(headers)
    for number, puzzle in enumerate(puzzles, start=1):
        q1, q2, q3 = (questions[index] for index in puzzle.ids)
        values = (q1.answer, q2.answer, q3.answer)
        worksheet.append(
            [
                number,
                puzzle.operator,
                number_text(q1.answer),
                q1.category,
                q1.text,
                number_text(q2.answer),
                q2.category,
                q2.text,
                number_text(q3.answer),
                q3.category,
                q3.text,
                formula(puzzle.operator, values),
                puzzle.level,
                puzzle.score,
            ]
        )
    style_sheet(worksheet)
    widths = {"A": 12, "B": 12, "C": 14, "D": 30, "E": 72, "F": 14, "G": 30, "H": 72, "I": 14, "J": 30, "K": 72, "L": 28, "M": 18, "N": 16}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def style_sheet(worksheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"


def write_workbook(
    output_path: Path,
    input_path: Path,
    questions: list[Question],
    main: list[Puzzle],
    race: list[Puzzle],
    used: set[int],
) -> None:
    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = "Puzzels"
    write_puzzle_sheet(main_sheet, "Puzzel #", main, questions)

    race_sheet = workbook.create_sheet("Puzzel Race")
    write_puzzle_sheet(race_sheet, "Race #", race, questions)

    overview = workbook.create_sheet("Overzicht")
    overview.append(["NETTO – hoofdset en Puzzel Race", None])
    overview.append(["Bronbestand", str(input_path)])
    overview.append(["Vragen ingelezen", len(questions)])
    overview.append(["Hoofd-puzzels", len(main)])
    overview.append(["Race-puzzels", len(race)])
    overview.append(["Totaal unieke gebruikte vragen", len(used)])
    overview.append(["Niet gebruikte vragen", len(questions) - len(used)])
    overview.append([None, None])
    overview.append(["Hoofdset per operator", "Aantal"])
    main_ops = Counter(puzzle.operator for puzzle in main)
    for operator in OPERATORS:
        overview.append([operator, main_ops[operator]])
    overview.append([None, None])
    overview.append(["Hoofdset per difficulty", "Aantal"])
    main_levels = Counter(puzzle.level for puzzle in main)
    for level in LEVELS:
        overview.append([level, main_levels[level]])
    overview.append([None, None])
    overview.append(["Race per operator", "Aantal"])
    race_ops = Counter(puzzle.operator for puzzle in race)
    for operator in OPERATORS:
        overview.append([operator, race_ops[operator]])
    overview.append([None, None])
    overview.append(["Race per difficulty", "Aantal"])
    race_levels = Counter(puzzle.level for puzzle in race)
    for level in LEVELS:
        overview.append([level, race_levels[level]])
    style_sheet(overview)
    overview.column_dimensions["A"].width = 38
    overview.column_dimensions["B"].width = 68

    unused_sheet = workbook.create_sheet("Niet gebruikt")
    unused_sheet.append(["Bronrij", "Categorie", "Vraag", "Antwoord"])
    for index, question in enumerate(questions):
        if index not in used:
            unused_sheet.append([question.source_row, question.category, question.text, number_text(question.answer)])
    style_sheet(unused_sheet)
    for column, width in {"A": 12, "B": 30, "C": 90, "D": 16}.items():
        unused_sheet.column_dimensions[column].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def verify(main: list[Puzzle], race: list[Puzzle], questions: list[Question], used: set[int]) -> None:
    assert len(main) == 200
    assert Counter(p.operator for p in main) == Counter({operator: 50 for operator in OPERATORS})
    assert Counter(p.level for p in main) == Counter({level: 50 for level in LEVELS})
    all_ids = [index for puzzle in main + race for index in puzzle.ids]
    assert len(all_ids) == len(set(all_ids)), "Een vraag is meer dan één keer gebruikt."
    for puzzle in main + race:
        q1, q2, q3 = (questions[index] for index in puzzle.ids)
        if puzzle.operator == "×":
            assert q1.answer * q2.answer == q3.answer
        elif puzzle.operator == "÷":
            assert q2.answer != 0 and q1.answer / q2.answer == q3.answer
        elif puzzle.operator == "+":
            assert q1.answer + q2.answer == q3.answer
        else:
            assert q1.answer - q2.answer == q3.answer
    assert len(used) == len(all_ids)


def main() -> None:
    parser = argparse.ArgumentParser(description="Maak Netto-hoofdpuzzels en Puzzel Race-puzzels.")
    parser.add_argument("--input", help="Pad naar de gecategoriseerde vragenbank.")
    parser.add_argument("--output", help="Pad naar de nieuwe puzzelbank.")
    parser.add_argument("--seed", type=int, default=20260831, help="Seed voor reproduceerbare selectie.")
    args = parser.parse_args()

    input_path = resolve_path(args.input, DEFAULT_INPUT)
    output_path = resolve_path(args.output, DEFAULT_OUTPUT)
    questions, skipped = load_questions(input_path)
    if len(questions) < 600:
        raise ValueError(f"Er zijn {len(questions)} bruikbare vragen; minimaal 600 nodig.")

    main_puzzles, main_used = choose_main_set(questions, args.seed)
    race_puzzles, race_used = choose_race_set(questions, main_used, args.seed + 1)
    used = main_used | race_used
    verify(main_puzzles, race_puzzles, questions, used)
    write_workbook(output_path, input_path, questions, main_puzzles, race_puzzles, used)

    print(f"Klaar: {len(main_puzzles)} hoofd-puzzels en {len(race_puzzles)} race-puzzels")
    print("Hoofdset operators:", dict(Counter(p.operator for p in main_puzzles)))
    print("Hoofdset difficulty:", dict(Counter(p.level for p in main_puzzles)))
    print(f"Unieke vragen gebruikt: {len(used)} van {len(questions)}")
    print(f"Niet gebruikt: {len(questions) - len(used)}")
    if skipped:
        print(f"Overgeslagen regels: {len(skipped)}")


if __name__ == "__main__":
    main()
