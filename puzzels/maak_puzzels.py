#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genereer Netto-puzzels uit de gecategoriseerde vragenbank.

Regels:
- Iedere vraag kan maximaal één keer worden gebruikt.
- Eerst worden vermenigvuldig- en deelsommen gemaakt.
- Vermenigvuldig- en deelsommen worden per ronde gebalanceerd.
- Daarna worden plus- en minsommen gemaakt, ook per ronde gebalanceerd.
- De formule moet exact kloppen.
- Het resultaat wordt naar een nieuw Excel-bestand geschreven.

Voorbeeld:
    python maak_puzzels.py

Optionele argumenten:
    python maak_puzzels.py --input "1000+ vragen netjes gecategoriseerd.xlsx"
    python maak_puzzels.py --output "netto_puzzels_nieuw.xlsx"
    python maak_puzzels.py --target-per-operator 250 --md-target 90
"""

from __future__ import annotations

import argparse
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from fractions import Fraction
from pathlib import Path
from typing import DefaultDict, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = "1000+ vragen netjes gecategoriseerd.xlsx"
DEFAULT_OUTPUT = "netto_puzzels_definitief.xlsx"
DEFAULT_MULTIPLICATION_DIVISION_TARGET = 90

OPERATORS = ("×", "÷", "+", "−")
OPERATOR_WEIGHTS = {"+": 0.0, "−": 8.0, "×": 14.0, "÷": 18.0}
MAX_IDS_PER_VALUE = 6
DEFAULT_MAX_CANDIDATES = 4000


@dataclass(frozen=True)
class Question:
    source_row: int
    category: str
    text: str
    answer: Fraction


@dataclass(frozen=True)
class Candidate:
    operator: str
    ids: tuple[int, int, int]
    priority: tuple[float, ...]



def parse_number(value: object) -> Fraction | None:
    """Zet een Excelwaarde veilig om naar een exacte Fraction."""
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, Fraction):
        return value
    if isinstance(value, int):
        return Fraction(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return Fraction(str(value))

    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        return None

    # Ondersteun zowel 12,5 als 12.5 en eenvoudige duizendtalsnotatie.
    if re.fullmatch(r"[-+]?\d{1,3}(?:\.\d{3})+", text):
        text = text.replace(".", "")
    elif "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return None

    try:
        return Fraction(text)
    except (ValueError, ZeroDivisionError):
        return None



def normalize_text(text: str) -> str:
    return " ".join(text.casefold().split())



def format_number(value: Fraction) -> str:
    """Maak een leesbare formulewaarde zonder duizendtalsnotatie."""
    if value.denominator == 1:
        return str(value.numerator)

    rendered = f"{float(value):.12f}".rstrip("0").rstrip(".")
    return rendered



def excel_number(value: Fraction) -> int | float:
    if value.denominator == 1:
        return value.numerator
    return float(value)



def resolve_path(argument: str | None, default_name: str) -> Path:
    if argument:
        return Path(argument).expanduser()
    return BASE_DIR / default_name



def load_questions(path: Path) -> tuple[list[Question], list[str]]:
    if not path.exists():
        raise FileNotFoundError(f"Invoerbestand niet gevonden: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Alle vragen" not in workbook.sheetnames:
        raise ValueError("Het invoerbestand bevat geen tabblad 'Alle vragen'.")

    worksheet = workbook["Alle vragen"]
    questions: list[Question] = []
    skipped: list[str] = []
    seen_texts: set[str] = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if len(row) < 3 or row[1] in (None, ""):
            continue

        text = str(row[1]).strip()
        answer = parse_number(row[2])
        if answer is None:
            skipped.append(
                f"Rij {row_number}: niet-numeriek antwoord overgeslagen ({row[2]!r})"
            )
            continue

        normalized = normalize_text(text)
        if normalized in seen_texts:
            skipped.append(f"Rij {row_number}: dubbele vraagtekst overgeslagen")
            continue

        seen_texts.add(normalized)
        questions.append(
            Question(
                source_row=row_number,
                category=str(row[0] or "Algemeen"),
                text=text,
                answer=answer,
            )
        )

    workbook.close()
    return questions, skipped



def build_buckets(
    questions: list[Question], used: set[int]
) -> DefaultDict[Fraction, list[int]]:
    buckets: DefaultDict[Fraction, list[int]] = defaultdict(list)
    for index, question in enumerate(questions):
        if index not in used:
            buckets[question.answer].append(index)
    return buckets



def distinct_id_combinations(
    first: list[int], second: list[int], third: list[int], limit: int = 4
) -> Iterable[tuple[int, int, int]]:
    """Geef enkele combinaties met drie verschillende vraag-ID's."""
    found = 0
    for first_id in first[:MAX_IDS_PER_VALUE]:
        for second_id in second[:MAX_IDS_PER_VALUE]:
            if second_id == first_id:
                continue
            for third_id in third[:MAX_IDS_PER_VALUE]:
                if third_id in (first_id, second_id):
                    continue
                yield first_id, second_id, third_id
                found += 1
                if found >= limit:
                    return



def candidate_priority(
    operator: str,
    values: tuple[Fraction, Fraction, Fraction],
    buckets: DefaultDict[Fraction, list[int]],
) -> tuple[float, ...]:
    scarcity = sum(1.0 / len(buckets[value]) for value in values)
    magnitude = sum(math.log10(abs(float(value)) + 1.0) for value in values)

    # Bij min wordt de grootste eerste waarde eerst geprobeerd, zoals gevraagd.
    if operator == "−":
        return scarcity, -float(values[0]), float(values[1]), -magnitude
    return scarcity, -magnitude



def find_candidates(
    operator: str,
    questions: list[Question],
    used: set[int],
    max_candidates: int = DEFAULT_MAX_CANDIDATES,
) -> list[Candidate]:
    """Vind mogelijke correcte triples uit de nog vrije vragen.

    Er worden alleen kandidaten met drie verschillende vraag-ID's teruggegeven.
    De lijst is bewust begrensd zodat ook grote vragenbanken snel blijven.
    """
    buckets = build_buckets(questions, used)
    values = sorted(buckets)
    collected: list[Candidate] = []
    seen: set[tuple[int, int, int]] = set()
    scan_cap = max_candidates * 4

    def add_candidates(
        ids: tuple[int, int, int],
        value_tuple: tuple[Fraction, Fraction, Fraction],
    ) -> None:
        if len(set(ids)) != 3:
            return

        candidate_ids = ids
        if operator in ("×", "+") and candidate_ids[0] > candidate_ids[1]:
            candidate_ids = (candidate_ids[1], candidate_ids[0], candidate_ids[2])

        if candidate_ids in seen:
            return
        seen.add(candidate_ids)
        collected.append(
            Candidate(
                operator=operator,
                ids=candidate_ids,
                priority=candidate_priority(operator, value_tuple, buckets),
            )
        )

    if operator == "×":
        for first_position, first_value in enumerate(values):
            if first_value <= 0 or first_value == 1:
                continue
            for second_value in values[first_position:]:
                if second_value <= 0 or second_value == 1:
                    continue
                result = first_value * second_value
                if result not in buckets:
                    continue
                for ids in distinct_id_combinations(
                    buckets[first_value],
                    buckets[second_value],
                    buckets[result],
                ):
                    add_candidates(
                        ids,
                        (first_value, second_value, result),
                    )
                    if len(collected) >= scan_cap:
                        break
                if len(collected) >= scan_cap:
                    break
            if len(collected) >= scan_cap:
                break

    elif operator == "÷":
        divisor_values = [value for value in values if value > 0 and value != 1]
        quotient_values = [value for value in values if value > 0 and value != 1]
        for divisor in divisor_values:
            for quotient in quotient_values:
                dividend = divisor * quotient
                if dividend not in buckets:
                    continue
                for ids in distinct_id_combinations(
                    buckets[dividend],
                    buckets[divisor],
                    buckets[quotient],
                ):
                    # q1 ÷ q2 = q3
                    add_candidates(ids, (dividend, divisor, quotient))
                    if len(collected) >= scan_cap:
                        break
                if len(collected) >= scan_cap:
                    break
            if len(collected) >= scan_cap:
                break

    elif operator == "+":
        positive_values = [value for value in values if value > 0]
        for first_position, first_value in enumerate(positive_values):
            for second_value in positive_values[first_position:]:
                result = first_value + second_value
                if result not in buckets:
                    continue
                for ids in distinct_id_combinations(
                    buckets[first_value],
                    buckets[second_value],
                    buckets[result],
                ):
                    add_candidates(ids, (first_value, second_value, result))
                    if len(collected) >= scan_cap:
                        break
                if len(collected) >= scan_cap:
                    break
            if len(collected) >= scan_cap:
                break

    elif operator == "−":
        first_values = sorted(
            (value for value in values if value > 0), reverse=True
        )
        second_values = sorted(value for value in values if value > 0)
        for first_value in first_values:
            for second_value in second_values:
                if second_value >= first_value:
                    continue
                result = first_value - second_value
                if result not in buckets or result <= 0:
                    continue
                for ids in distinct_id_combinations(
                    buckets[first_value],
                    buckets[second_value],
                    buckets[result],
                ):
                    # q1 − q2 = q3; eerste getal is bewust groter.
                    add_candidates(ids, (first_value, second_value, result))
                    if len(collected) >= scan_cap:
                        break
                if len(collected) >= scan_cap:
                    break
            if len(collected) >= scan_cap:
                break

    collected.sort(key=lambda candidate: candidate.priority)
    return collected[:max_candidates]



def choose_compatible(
    first_candidates: list[Candidate], second_candidates: list[Candidate]
) -> tuple[Candidate, Candidate] | None:
    """Kies twee kandidaten die geen vraag delen."""
    for first in first_candidates:
        first_ids = set(first.ids)
        for second in second_candidates:
            if first_ids.isdisjoint(second.ids):
                return first, second
    return None



def generate_balanced_phase(
    first_operator: str,
    second_operator: str,
    questions: list[Question],
    used: set[int],
    output_candidates: list[Candidate],
    target_per_operator: int,
    max_candidates: int,
) -> int:
    """Maak gelijke aantallen van twee operatoren totdat dat niet meer kan.

    Kandidaten worden één keer opgebouwd per fase. De vorige implementatie
    bouwde de volledige kandidaatlijst na iedere ronde opnieuw op; met een
    vragenbank van bijna duizend regels kon dat onnodig lang duren.
    """
    first_candidates = find_candidates(
        first_operator, questions, used, max_candidates
    )
    second_candidates = find_candidates(
        second_operator, questions, used, max_candidates
    )
    pair_count = 0

    while pair_count < target_per_operator:
        first = next(
            (candidate for candidate in first_candidates
             if set(candidate.ids).isdisjoint(used)),
            None,
        )
        if first is None:
            break

        first_ids = set(first.ids)
        second = next(
            (candidate for candidate in second_candidates
             if first_ids.isdisjoint(candidate.ids)
             and set(candidate.ids).isdisjoint(used)),
            None,
        )
        if second is None:
            # Deze fase kan met de resterende kandidaten niet verder worden
            # gebalanceerd; de volgende operatorfase mag daarna wel doorgaan.
            break

        output_candidates.extend((first, second))
        used.update(first.ids)
        used.update(second.ids)
        pair_count += 1

    return pair_count



def _relation_values(operator: str, value_set: set[Fraction]) -> list[tuple[Fraction, Fraction, Fraction]]:
    """Maak unieke waarderelaties; vraag-ID's worden pas bij allocatie gekozen."""
    values = sorted(value_set)
    relations: list[tuple[Fraction, Fraction, Fraction]] = []

    if operator == "×":
        usable = [value for value in values if value > 0 and value != 1]
        for first_position, first in enumerate(usable):
            for second in usable[first_position:]:
                result = first * second
                if result in value_set:
                    relations.append((first, second, result))
    elif operator == "÷":
        usable = [value for value in values if value > 0 and value != 1]
        for divisor in usable:
            for quotient in usable:
                dividend = divisor * quotient
                if dividend in value_set:
                    relations.append((dividend, divisor, quotient))
    elif operator == "+":
        usable = [value for value in values if value > 0]
        for first_position, first in enumerate(usable):
            for second in usable[first_position:]:
                result = first + second
                if result in value_set:
                    relations.append((first, second, result))
    elif operator == "−":
        positive = [value for value in values if value > 0]
        for first in positive:
            for second in positive:
                if second >= first:
                    continue
                result = first - second
                if result > 0 and result in value_set:
                    relations.append((first, second, result))

    return relations



def _relation_counts(relation: tuple[Fraction, Fraction, Fraction]) -> Counter:
    return Counter(relation)



def _fits_capacity(
    relation: tuple[Fraction, Fraction, Fraction], capacities: dict[Fraction, int]
) -> bool:
    return all(count <= capacities.get(value, 0) for value, count in _relation_counts(relation).items())



def _relation_priority(
    relation: tuple[Fraction, Fraction, Fraction],
    capacities: dict[Fraction, int],
    degrees: Counter,
) -> tuple[float, float, float]:
    counts = _relation_counts(relation)
    # Gebruik eerst waarden met weinig beschikbare vragen of weinig mogelijke
    # relaties. Zo blijven veel algemene waarden beschikbaar als reserve.
    scarcity = sum(count / max(1, capacities.get(value, 0)) for value, count in counts.items())
    rarity = sum(count / max(1, degrees.get(value, 0)) for value, count in counts.items())
    magnitude = sum(math.log10(abs(float(value)) + 1.0) for value in relation)
    return (scarcity, rarity, -magnitude)



def _allocate_relation(
    relation: tuple[Fraction, Fraction, Fraction],
    available: DefaultDict[Fraction, list[int]],
) -> tuple[int, int, int]:
    ids: list[int] = []
    for value in relation:
        ids.append(available[value].pop())
    return ids[0], ids[1], ids[2]



def generate_balanced_phase_by_value(
    first_operator: str,
    second_operator: str,
    questions: list[Question],
    used: set[int],
    output_candidates: list[Candidate],
    target_per_operator: int,
) -> int:
    """Maak een gebalanceerde operatorfase zonder vragen opnieuw te gebruiken.

    Deze waardegerichte variant voorkomt dat meerdere vragen met hetzelfde
    antwoord elkaar onnodig blokkeren tijdens het zoeken naar kandidaten.
    """
    available: DefaultDict[Fraction, list[int]] = defaultdict(list)
    for index, question in enumerate(questions):
        if index not in used:
            available[question.answer].append(index)

    value_set = set(available)
    first_relations = _relation_values(first_operator, value_set)
    second_relations = _relation_values(second_operator, value_set)

    def degrees(relations: list[tuple[Fraction, Fraction, Fraction]]) -> Counter:
        result: Counter = Counter()
        for relation in relations:
            result.update(set(relation))
        return result

    first_degrees = degrees(first_relations)
    second_degrees = degrees(second_relations)
    pairs = 0

    while pairs < target_per_operator:
        capacities = {value: len(ids) for value, ids in available.items()}
        valid_first = [
            relation for relation in first_relations
            if _fits_capacity(relation, capacities)
        ]
        valid_second = [
            relation for relation in second_relations
            if _fits_capacity(relation, capacities)
        ]
        if not valid_first or not valid_second:
            break

        valid_first.sort(key=lambda r: _relation_priority(r, capacities, first_degrees))
        valid_second.sort(key=lambda r: _relation_priority(r, capacities, second_degrees))

        chosen: tuple[tuple[Fraction, Fraction, Fraction], tuple[Fraction, Fraction, Fraction]] | None = None
        # In de praktijk past een vroege zeldzame relatie bijna altijd bij een
        # vroege relatie van de andere operator. De bredere fallback voorkomt
        # dat een slechte eerste keuze de hele fase vroegtijdig stopt.
        for first_relation in valid_first[:2000]:
            first_counts = _relation_counts(first_relation)
            for second_relation in valid_second[:2000]:
                total = first_counts + _relation_counts(second_relation)
                if all(count <= capacities.get(value, 0) for value, count in total.items()):
                    chosen = first_relation, second_relation
                    break
            if chosen is not None:
                break

        if chosen is None:
            break

        first_relation, second_relation = chosen
        first_ids = _allocate_relation(first_relation, available)
        second_ids = _allocate_relation(second_relation, available)
        first_candidate = Candidate(first_operator, first_ids, (0.0,))
        second_candidate = Candidate(second_operator, second_ids, (0.0,))
        output_candidates.extend((first_candidate, second_candidate))
        used.update(first_ids)
        used.update(second_ids)
        pairs += 1

    return pairs



def difficulty_for(
    operator: str, answers: tuple[Fraction, Fraction, Fraction]
) -> tuple[str, float]:
    score = OPERATOR_WEIGHTS[operator] + sum(
        math.log10(abs(float(answer)) + 1.0) * 10.0 for answer in answers
    )
    score = round(score, 2)

    if score < 35:
        level = "Easy"
    elif score < 60:
        level = "Intermediate"
    elif score < 90:
        level = "Hard"
    else:
        level = "Extremely Hard"
    return level, score



def formula_for(
    operator: str, answers: tuple[Fraction, Fraction, Fraction]
) -> str:
    return " {} ".format(operator).join(
        (format_number(answers[0]), format_number(answers[1]))
    ) + f" = {format_number(answers[2])}"



def write_workbook(
    output_path: Path,
    source_path: Path,
    questions: list[Question],
    used: set[int],
    candidates: list[Candidate],
) -> None:
    workbook = Workbook()
    puzzle_sheet = workbook.active
    puzzle_sheet.title = "Puzzels"

    headers = [
        "Puzzel #",
        "Bewerking",
        "Antwoord 1",
        "Categorie 1",
        "Vraag 1",
        "Antwoord 2",
        "Categorie 2",
        "Vraag 2",
        "Antwoord 3",
        "Categorie 3",
        "Vraag 3",
        "Exacte formule",
        "Difficulty",
        "Difficulty score",
    ]
    puzzle_sheet.append(headers)

    for puzzle_number, candidate in enumerate(candidates, start=1):
        q1, q2, q3 = (questions[index] for index in candidate.ids)
        answers = (q1.answer, q2.answer, q3.answer)
        difficulty, difficulty_score = difficulty_for(candidate.operator, answers)

        # Controleer de rekenregel nogmaals voordat de rij wordt opgeslagen.
        if candidate.operator == "×":
            assert q1.answer * q2.answer == q3.answer
        elif candidate.operator == "÷":
            assert q2.answer != 0 and q1.answer / q2.answer == q3.answer
        elif candidate.operator == "+":
            assert q1.answer + q2.answer == q3.answer
        elif candidate.operator == "−":
            assert q1.answer - q2.answer == q3.answer

        puzzle_sheet.append(
            [
                puzzle_number,
                candidate.operator,
                excel_number(q1.answer),
                q1.category,
                q1.text,
                excel_number(q2.answer),
                q2.category,
                q2.text,
                excel_number(q3.answer),
                q3.category,
                q3.text,
                formula_for(candidate.operator, answers),
                difficulty,
                difficulty_score,
            ]
        )

    style_sheet(
        puzzle_sheet,
        widths={
            "A": 11,
            "B": 12,
            "C": 14,
            "D": 28,
            "E": 72,
            "F": 14,
            "G": 28,
            "H": 72,
            "I": 14,
            "J": 28,
            "K": 72,
            "L": 28,
            "M": 18,
            "N": 16,
        },
    )

    overview = workbook.create_sheet("Overzicht")
    overview.append(["NETTO – puzzelgenerator", None])
    overview.append(["Bronbestand", str(source_path)])
    overview.append(["Vragen ingelezen", len(questions)])
    overview.append(["Vragen gebruikt", len(used)])
    overview.append(["Vragen niet gebruikt", len(questions) - len(used)])
    overview.append(["Totaal puzzels", len(candidates)])

    counts = Counter(candidate.operator for candidate in candidates)
    overview.append(["Vermenigvuldigen (×)", counts["×"]])
    overview.append(["Delen (÷)", counts["÷"]])
    overview.append(["Plus (+)", counts["+"]])
    overview.append(["Min (−)", counts["−"]])
    overview.append(["× en ÷ evenveel", counts["×"] == counts["÷"]])
    overview.append(["+ en − evenveel", counts["+"] == counts["−"]])

    difficulty_counts = Counter()
    for candidate in candidates:
        question_answers = tuple(questions[index].answer for index in candidate.ids)
        difficulty_counts[difficulty_for(candidate.operator, question_answers)[0]] += 1

    overview.append([None, None])
    overview.append(["Difficulty", "Aantal"])
    for level in ("Easy", "Intermediate", "Hard", "Extremely Hard"):
        overview.append([level, difficulty_counts[level]])

    style_sheet(overview, widths={"A": 34, "B": 64})

    unused_sheet = workbook.create_sheet("Niet gebruikt")
    unused_sheet.append(["Bronrij", "Categorie", "Vraag", "Antwoord"])
    for index, question in enumerate(questions):
        if index not in used:
            unused_sheet.append(
                [
                    question.source_row,
                    question.category,
                    question.text,
                    excel_number(question.answer),
                ]
            )
    style_sheet(
        unused_sheet,
        widths={"A": 12, "B": 30, "C": 90, "D": 16},
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)



def style_sheet(worksheet, widths: dict[str, float]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        )



def main() -> None:
    parser = argparse.ArgumentParser(description="Maak unieke Netto-rekenpuzzels.")
    parser.add_argument("--input", help="Pad naar de vragenbank.")
    parser.add_argument("--output", help="Pad voor het nieuwe Excel-bestand.")
    parser.add_argument(
        "--target-per-operator",
        type=int,
        default=250,
        help="Maximaal aantal puzzels per operator (standaard: 250).",
    )
    parser.add_argument(
        "--md-target",
        type=int,
        default=DEFAULT_MULTIPLICATION_DIVISION_TARGET,
        help="Maximaal aantal keer/deel-paren; deze fase krijgt prioriteit.",
    )
    parser.add_argument(
        "--max-candidates",
        type=int,
        default=DEFAULT_MAX_CANDIDATES,
        help="Aantal kandidaatcombinaties per zoekronde.",
    )
    args = parser.parse_args()

    if args.target_per_operator < 0:
        parser.error("--target-per-operator moet minimaal 0 zijn.")
    if args.md_target < 0:
        parser.error("--md-target moet minimaal 0 zijn.")
    if args.max_candidates < 10:
        parser.error("--max-candidates moet minimaal 10 zijn.")

    input_path = resolve_path(args.input, DEFAULT_INPUT)
    output_path = resolve_path(args.output, DEFAULT_OUTPUT)
    questions, skipped = load_questions(input_path)

    if len(questions) < 3:
        raise ValueError("Er zijn minder dan drie bruikbare vragen gevonden.")

    used: set[int] = set()
    candidates: list[Candidate] = []

    multiplication_division_rounds = generate_balanced_phase_by_value(
        "×",
        "÷",
        questions,
        used,
        candidates,
        min(args.target_per_operator, args.md_target),
    )
    plus_minus_rounds = generate_balanced_phase_by_value(
        "+",
        "−",
        questions,
        used,
        candidates,
        args.target_per_operator,
    )

    write_workbook(output_path, input_path, questions, used, candidates)

    counts = Counter(candidate.operator for candidate in candidates)
    print(f"Klaar: {len(candidates)} puzzels geschreven naar {output_path}")
    print(
        "Operatoren: "
        f"× {counts['×']}, ÷ {counts['÷']}, + {counts['+']}, − {counts['−']}"
    )
    print(f"Gebruikte vragen: {len(used)} van {len(questions)}")
    print(f"Niet gebruikte vragen: {len(questions) - len(used)}")
    print(f"×/÷-rondes: {multiplication_division_rounds}")
    print(f"+/-rondes: {plus_minus_rounds}")
    arithmetic_puzzles = len(candidates)
    plus_minus_share = ((counts['+'] + counts['−']) / arithmetic_puzzles) if arithmetic_puzzles else 0
    print(f"+/- aandeel: {plus_minus_share:.1%}")
    if arithmetic_puzzles and plus_minus_share < 1 / 3:
        raise RuntimeError("De gegenereerde set voldoet niet aan minimaal een derde +/−-puzzels.")

    if skipped:
        print(f"Overgeslagen invoerregels: {len(skipped)}")
        for message in skipped[:10]:
            print(f"  - {message}")
        if len(skipped) > 10:
            print(f"  ... en nog {len(skipped) - 10}")


if __name__ == "__main__":
    main()
