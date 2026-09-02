#!/usr/bin/env python3
"""Hervatbare review-helper voor de Nederlandse/Engelse vragenbank.

Dit script automatiseert alleen veilige administratieve stappen. Het verzint geen
bronnen en keurt geen inhoud goed zonder menselijke/broncontrole.
"""
from __future__ import annotations

import argparse
import re
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

URL_RE = re.compile(r"https?://[^\s]+", re.I)


def normalize_url(value):
    if not value:
        return ""
    match = URL_RE.search(str(value))
    return match.group(0).rstrip(".,);]") if match else ""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="vragen/1000+ vragen netjes gecategoriseerd.xlsx")
    parser.add_argument("--output", default=None)
    parser.add_argument("--start", type=int, default=2)
    parser.add_argument("--limit", type=int, default=0, help="0 = alle resterende rijen")
    args = parser.parse_args()

    source = Path(args.input)
    output = Path(args.output) if args.output else source
    wb = load_workbook(source)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    required = ["Vraag", "Categorie", "Antwoord", "Eenheid", "Categorie EN", "Question EN", "Answer EN", "Source EN", "Source checked EN", "Review EN"]
    missing = [name for name in required if name not in col]
    if missing:
        raise SystemExit(f"Ontbrekende kolommen: {', '.join(missing)}")

    processed = 0
    for row in range(max(2, args.start), ws.max_row + 1):
        if args.limit and processed >= args.limit:
            break
        question = ws.cell(row, col["Vraag"]).value
        if not question or ws.cell(row, col["Review EN"]).value not in (None, ""):
            continue

        category = ws.cell(row, col["Categorie"]).value or "Algemene kennis"
        answer = ws.cell(row, col["Antwoord"]).value
        unit = ws.cell(row, col["Eenheid"]).value
        ws.cell(row, col["Categorie EN"]).value = category
        ws.cell(row, col["Question EN"]).value = str(question)
        ws.cell(row, col["Answer EN"]).value = f"{answer} {unit}".strip() if answer is not None else ""

        # Alleen URL's behouden; een lege of niet-passende bron blijft expliciet review nodig.
        current_url = normalize_url(ws.cell(row, col["Source EN"]).value)
        ws.cell(row, col["Source EN"]).value = current_url or None
        ws.cell(row, col["Source checked EN"]).value = str(date.today())
        ws.cell(row, col["Review EN"]).value = 2
        note_cell = ws.cell(row, col["Opmerking"]) if "Opmerking" in col else None
        if note_cell is not None and not note_cell.value:
            note_cell.value = "Manual English/source verification required."
        processed += 1
        if processed % 25 == 0:
            wb.save(output)

    wb.save(output)
    print(f"Saved: {output}")
    print(f"Rows administratively prepared this run: {processed}")
    print("Review EN blijft 2 totdat de Engelse tekst en de specifieke bron inhoudelijk zijn gecontroleerd.")


if __name__ == "__main__":
    main()
